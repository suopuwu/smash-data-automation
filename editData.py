import openpyxl
import re
sheetName = input('sheet name: ')
wb = openpyxl.load_workbook(sheetName)
charactersText = open('characters.txt')
characterList = charactersText.read()
characterList = characterList.split(', ')
if 'GlossaryNotes' in wb.sheetnames:
    del wb['GlossaryNotes']
else:
    print('GlossaryNotes either doesn\'t exist or is named weirdly')
counter = 0
for character in wb.sheetnames:
    wb[character]['A1'] = 'name'
    wb[character].title = re.sub('.* -' , '', character).strip()
    character = re.sub('.* -' , '', character).strip()
    #we redefine character to its new value to make the line below more readable
    wb[character].title = 'temp'
    #if you try to set the title to itself but lowercase, it'll append a 1 because titles are not case sensitive. Thus, we must set a temporary value first.
    wb['temp'].title = character.lower()
    counter += 1
print(str(counter) + ' characters edited')

#generates a list of characters that should exist in the spreadsheet, but don't
missingCharacters = []
for character in characterList:
    if character not in wb.sheetnames:
        missingCharacters.append(character)

#generates a list of characters that exist in the spreadsheet, but don't have a matching name in the character list
unmatchedCharacters = []
for character in wb.sheetnames:
    if character not in characterList:
        unmatchedCharacters.append(character)

print(str(len(missingCharacters)) + ' characters missing')

for character in missingCharacters:
    print('\n\n\n\n'+ character + ' is missing, choose a character that matches up')
    choices = []
    for unmatchedCharacter in unmatchedCharacters:
        choices.append(unmatchedCharacter + '[' + str(unmatchedCharacters.index(unmatchedCharacter)) + ']')
    print(' '.join(choices))
    choice = input('Type a number: ')
    source = wb[unmatchedCharacters[int(choice)]]
    temp = wb.copy_worksheet(source)
    temp.title = character

for character in unmatchedCharacters:
    del wb[character]

missingCharacters = []
for character in characterList:
    if character not in wb.sheetnames:
        missingCharacters.append(character)
print('Missing characters: ' + str(len(missingCharacters)))
print('saving...')
wb.save(sheetName)
input('done! press enter to exit.')
